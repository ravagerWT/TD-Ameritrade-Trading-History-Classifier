import os
import time
import datetime
import shutil
import PySimpleGUI as sg
from openpyxl import load_workbook, utils
from openpyxl.styles import PatternFill, Alignment
# import openpyxl.worksheet
import ctypes
from datetime import datetime, date, time
import re
import json
import settings
import language

# openpyxl.utils.cell.coordinate_from_string('B3')  // ('B', 3)
# openpyxl.utils.cell.column_index_from_string(a[0])  // 2
# openpyxl.utils.cell.coordinate_to_tuple('B3')  // (3, 2)
# openpyxl.utils.cell.get_column_letter(3) // C

program_ver = 'Dev 1.0'

#// TODO:實作介面語言字串全域變數
# load program setting from settings.json
def loadSetting(setting_file_name='settings.json'):
    """[load program setting from settings.json]
    
    Keyword Arguments:
        setting_file_name {str} -- [setting file name] (default: {'settings.json'})
    
    Returns:
        [type] -- [description]
    """
    try:
        # open json file and read
        with open(setting_file_name, 'r', encoding="utf-8") as setting:
            my_setting = json.loads(setting.read())
    except FileNotFoundError:
        MessageBox(None, setting_file_name + ' ' + lang.msg_box_file_not_exist, lang.msg_box_file_op_title, 0)
    else:
        return settings.Thfset(my_setting)
    finally:
        setting.close()

# edit program setting in settings.json
def editSetting(st, lang):
    odd_color_status = False
    even_color_status = False
    layout = [
        [sg.Text('Localization: '), sg.InputCombo(st.gen_ava_lang_for_GUI, size=(
            20, 1), default_value=st.gen_set_lang, key='set_lang', readonly=True)],
        [sg.Text('_' * 100, size=(55, 1))],
        [sg.Text('Excel format setting: ')],
        [sg.Text('Odd column color (Hex):', size=(18, 1)), sg.InputText('FFF2CC', key='odd_col_color')],
        [sg.Text('Even column color (Hex):', size=(18, 1)), sg.InputText('E2EFDA', key='even_col_color')],
        [sg.Text('Display date format:', size=(18, 1)), sg.InputText(st.xls_fmt_display_date_format, key='date_fmt')],
        [sg.Button('OK'), sg.Cancel('Cancel')]
    ]
    
    window = sg.Window('Program Settings', auto_size_text=True,
                   default_element_size=(40, 10)).Layout(layout)

    while True:
        event, values = window.Read()
        print('event: ', event, '\nvalues:', values)  # debug message
        if event == 'OK':
            # https://stackoverflow.com/questions/30241375/python-how-to-check-if-string-is-a-hex-color-code
            if re.search(r'^(?:[0-9a-fA-F]{3}){1,2}$', values['odd_col_color']):
                st.gen_set_lang = values['odd_col_color']
                odd_color_status = True
            else:
                window.Element('odd_col_color').Update('Wrong')
                MessageBox(None, 'Odd column color format is worng',
                           'Color Format Wrong', 0)
            if re.search(r'^(?:[0-9a-fA-F]{3}){1,2}$', values['even_col_color']):
                st.gen_set_lang = values['even_col_color']
                even_color_status = True
            else:
                window.Element('even_col_color').Update('Wrong')
                MessageBox(None, 'Even column color format is worng',
                           'Color Format Wrong', 0)
            if odd_color_status and even_color_status:
                # // TODO:待實作設定檔儲存功能
                saveSetting(st)
                break
        elif event is None or event == 'Cancel':
            break

    window.close()

# save setting to settings.json
def saveSetting(settings):
    pass

# load langage from lang.json
def loadLang(lang_code='English (enUS)'):
    lang_file_name = 'lang_' + st.gen_set_lang[-5:-1] + '.json'
    try:
        # open json file and read
        with open(lang_file_name, 'r', encoding="utf-8") as lang:
            lang_setting = json.loads(lang.read())
    except FileNotFoundError:
        MessageBox(None, lang_file_name + ' ' + lang.msg_box_file_not_exist, lang.msg_box_file_op_title, 0)
    else:
        return language.Lang(lang_setting)
    finally:
        lang.close()

# setup window layout
def setWindow(lang, st):    
    #// TODO: 簡化或整合檔案載入及處理GUI介面
    # setup window layout
    layout = [[sg.Text(lang.gui_program_setting + ':'), sg.Text('', size=(20, 1), key='settings status')],
              [sg.FileBrowse(lang.gui_load_setting_file, file_types=((lang.gui_settings_file, "settings.json"),)),
               sg.Button(lang.gui_apply_settings, key='Apply Settings'), sg.Button(lang.gui_open_setting_editor, key='Open Setting Editor')],
              [sg.Text('_' * 100, size=(70, 1))],
              [sg.Text(lang.gui_load_trade_history_file + ':')],
              [sg.Text(lang.gui_file + ':', justification='right'),
               sg.Text('', size=(65, 1), key='it_filePath')],
              [sg.FileBrowse(file_types=((lang.gui_spreadsheet_files, "*.xls"),
                                         (lang.gui_spreadsheet_files, "*.xlsx"),), target='it_filePath'),
               sg.Button(lang.gui_process_history, key='Process History'), sg.Checkbox(
                   lang.gui_rem_last_xls_file, enable_events=True, key='Last xls chkbox'),
               sg.Checkbox(lang.gui_exp_error_log, default=st.gen_exp_error_log, enable_events=True, key='exp error log')],
              [sg.Text(lang.gui_result + ':'),
               sg.Text('', size=(20, 1), key='Result')],
              [sg.Text('_' * 100, size=(70, 1))],
              [sg.Text(lang.gui_ver + ': ' + program_ver, size=(80, 1), font='Arial 8'),
               sg.Button(lang.gui_exit, size=(5, 1), key='Exit')]]

    # rendering window
    window = sg.Window(lang.gui_title, auto_size_text=True, default_element_size=(
        40, 10), resizable=False).Layout(layout)
    return window

# get excel file name to be processed
def getXlsFileName(filePath, lang):
    if filePath == None or filePath == '':        
        MessageBox(None, lang.msg_box_select_file_first, lang.msg_box_file_op_title, 0)
        return 'PathError'                
    else:
        xls_fileName = os.path.basename(filePath)        
        os.chdir(os.path.dirname(filePath))    
        return xls_fileName

# excel processor
def excelProcessor(xls_fileName, exp_error_log, symbol_list = []):
    sheet_list = ['Sorted trade history','ORDINARY DIVIDEND','W-8 WITHHOLDING','WIRING INFO','Ver','log']    
    error_log_qty = 0
    if symbol_list == []: gather_symbol = True
    else: gather_symbol = False
    # loading workbook
    wb = load_workbook(xls_fileName)
    # create sheets
    for i in range(len(sheet_list)):
        if not sheet_list[i] in wb.sheetnames:
            wb.create_sheet(sheet_list[i])
    
    ws_tran = wb["transactions"]
    ws_STH = wb["Sorted trade history"]
    ws_OD = wb["ORDINARY DIVIDEND"]
    ws_W8 = wb["W-8 WITHHOLDING"]
    ws_WI = wb["WIRING INFO"]
    ws_ver = wb["Ver"]
    ws_log = wb["log"]

    # setting layout
    ws_STH['A2'] = 'DATE' # Sorted trade history
    ws_OD['A1'] = 'DATE' # ORDINARY DIVIDEND
    ws_W8['A1'] = 'DATE' # W-8 WITHHOLDING
    ws_WI['A1'] = 'DATE' # WIRING INFO
    ws_WI['B1'] = 'Amount'
    ws_ver['A1'] = 'DATE' # Ver
    ws_ver['B1'] = 'Ver'
    ws_log['A1'] = 'Event'
    ws_log['B1'] = 'Message'

    # start sheets process
    iter_date_STH = ''
    iter_date_OD = ''
    iter_date_W8 = ''
    iter_date_WI = ''
    for i in range(2, ws_tran.max_row):
        tr_date = ws_tran.cell(i, 1)
        # date process
        temp_date_for_sheet = datetime.strptime(tr_date.value, "%m/%d/%Y")
        date_for_sheet = temp_date_for_sheet.strftime("%Y/%m/%d")
        tr_description = ws_tran.cell(i, 3)
        tr_qty = ws_tran.cell(i, 4)
        tr_symbol = ws_tran.cell(i, 5)        
        tr_price = ws_tran.cell(i, 6)
        tr_fee = ws_tran.cell(i, 7)
        tr_amount = ws_tran.cell(i, 8)
        # processing sheets format by stock symbols and gather all stock symbol from 'E5'
        if gather_symbol:
            if not tr_symbol.value in symbol_list and tr_symbol.value != None:
                symbol_list.append(tr_symbol.value)  # gather stock symbol
                symbol_index = symbol_list.index(tr_symbol.value)  # stock symbol
                ws_STH.cell(1, symbol_index*4+2).value = tr_symbol.value
                ws_STH.merge_cells(start_row=1, start_column=symbol_index *
                                4+2, end_row=1, end_column=symbol_index*4+5)  # merge cell
                ws_STH.cell(1, symbol_index*4+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                ws_STH.cell(2, symbol_index*4+2).value = 'Quantity'
                ws_STH.cell(2, symbol_index*4+3).value = 'Price'
                ws_STH.cell(2, symbol_index*4+4).value = 'Fee'
                ws_STH.cell(2, symbol_index*4+5).value = 'Amount'
                ws_OD.cell(1, symbol_index+2).value = tr_symbol.value
                ws_OD.cell(1, symbol_index+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                ws_W8.cell(1, symbol_index+2).value = tr_symbol.value
                ws_W8.cell(1, symbol_index+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                # print(tr_symbol.value)  # debug message
        else:
            for i in range(len(symbol_list)):
                symbol_index = symbol_list.index(tr_symbol.value)  # stock symbol
                ws_STH.cell(1, symbol_index*4+2).value = tr_symbol.value
                ws_STH.merge_cells(start_row=1, start_column=symbol_index *
                                4+2, end_row=1, end_column=symbol_index*4+5)  # merge cell
                ws_STH.cell(1, symbol_index*4+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                ws_STH.cell(2, symbol_index*4+2).value = 'Quantity'
                ws_STH.cell(2, symbol_index*4+3).value = 'Price'
                ws_STH.cell(2, symbol_index*4+4).value = 'Fee'
                ws_STH.cell(2, symbol_index*4+5).value = 'Amount'
                ws_OD.cell(1, symbol_index+2).value = tr_symbol.value
                ws_OD.cell(1, symbol_index+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                ws_W8.cell(1, symbol_index+2).value = tr_symbol.value
                ws_W8.cell(1, symbol_index+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                # print(tr_symbol.value)  # debug message

        # sorting trade history
        if 'WIRE INCOMING' in tr_description.value:
            if tr_date.value != iter_date_WI:
                ws_WI.insert_rows(2)  # add new row
                ws_WI.cell(2, 1).value = date_for_sheet  # date
                ws_WI.cell(2, 2).value = tr_amount.value  # amount
        elif 'Bought' in tr_description.value:            
            if tr_symbol.value in symbol_list:
                symbol_index = symbol_list.index(tr_symbol.value) # get index value in list
                if tr_date.value != iter_date_STH:
                    ws_STH.insert_rows(3)  # add new row                           
                    ws_STH.cell(3, 1).value = date_for_sheet # date     
                    iter_date_STH = tr_date.value
                ws_STH.cell(3, symbol_index*4+2).value = tr_qty.value
                ws_STH.cell(3, symbol_index*4+3).value = tr_price.value
                ws_STH.cell(3, symbol_index*4+4).value = tr_fee.value
                ws_STH.cell(3, symbol_index*4+5).value = tr_amount.value
            else: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = 'transaction symbol missing'
                    ws_log.cell(2, 2).value = 'not in symbol list: ' + tr_symbol.value + ' on '+ str(i) + 'th row'
                    error_log_qty += 1
        #// TODO: 待確認關鍵字
        elif 'Sold' in tr_description.value:
            ws_STH.insert_rows(3)  # add new row
            pass
        elif 'ORDINARY DIVIDEND' in tr_description.value:
            symbol_index = symbol_list.index(tr_symbol.value) # get index value in list
            if tr_date.value != iter_date_OD:
                ws_OD.insert_rows(2)  # add new row                           
                ws_OD.cell(2, 1).value = date_for_sheet # date     
                iter_date_OD = tr_date.value
            ws_OD.cell(2, symbol_index+2).value = tr_amount.value
        elif 'WITHHOLDING' in tr_description.value:            
            if tr_symbol.value == None: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = 'WITHHOLDING'
                    ws_log.cell(2, 2).value = 'No symbol information on ' + str(i) + 'th row'
                    error_log_qty += 1
                    # ws_W8.cell(2, len(symbol_list)+2).value = tr_amount.value
            else:
                if tr_date.value != iter_date_W8:
                    ws_W8.insert_rows(2)  # add new row
                    ws_W8.cell(2, 1).value = date_for_sheet  # date
                    iter_date_W8 = tr_date.value
                ws_W8.cell(2, symbol_index+2).value = tr_amount.value
        #// TODO: 待確認以下關鍵字：出金
        else: # export error message
            if exp_error_log:
                ws_log.insert_rows(2)
                ws_log.cell(2, 1).value = 'Description keyword missing'
                ws_log.cell(2, 2).value = 'not in the known keyword: ' + tr_description.value + ' on '+ str(i) + 'th row'
                error_log_qty += 1
                # print('not in the known keyword: ' + ws_tran.cell(i, 3).value)

    # version control
    file_version = ws_ver['B2'].value  # get current file version
    [file, ext] = os.path.splitext(xls_fileName)
    if file_version == None:
        ws_ver['A2'] = date.today().strftime("%Y/%m/%d")  # date
        ws_ver['B2'] = 0
        file_version = 0
    else:
        ws_ver.insert_rows(2)  # add new row
        ws_ver['A2'] = date.today().strftime("%Y/%m/%d")  # date
        file_version += 1  # update version number
        ws_ver['B2'] = file_version
    fileNameRev = file + '_r' + str(file_version) + ext
    
    # setting cell color
    for k in range(len(symbol_list)):
        color1 = "FFF2CC"
        color2 = "E2EFDA"
        if k % 2 == 0:
            color_fill = color2
        else:
            color_fill = color1
        for row in ws_STH.iter_rows(min_col=k*4+2, min_row=1, max_col=k*4+5, max_row=ws_STH.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor=color_fill, fill_type="solid")
        for row in ws_OD.iter_rows(min_col=k+2, min_row=1, max_col=k+2, max_row=ws_OD.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor=color_fill, fill_type="solid")
        for row in ws_W8.iter_rows(min_col=k+2, min_row=1, max_col=k+2, max_row=ws_W8.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor=color_fill, fill_type="solid")

        # ws_OD.column_dimensions[utils.cell.get_column_letter(
            # k+2)].fill = PatternFill(fgColor=color_fill, fill_type="solid")
        # ws_W8.column_dimensions[utils.cell.get_column_letter(
            # k+2)].fill = PatternFill(fgColor=color_fill, fill_type="solid")

    wb.save(fileNameRev)  # save processed file
    return error_log_qty
    
# Main Program
def main(window, lang):
    xls_fileName = None
    while True:
        event, values = window.Read()
        print('event: ', event, '\nvalues:', values)  # debug message
        if event == 'Apply Settings':            
            if values['Load Setting File'] == None or values['Load Setting File'] == '':
                MessageBox(None, lang.msg_box_select_file_first, lang.msg_box_file_op_title, 0)
            else:                                
                window.Close()
                return True
        elif event == 'Open Setting Editor':
            #// TODO:實作設定檔編輯功能
            editSetting(st, lang)
        elif event == 'Process History':
            #// TODO:待實作檔案名稱檢查
            xls_fileName = getXlsFileName(values['Browse'], lang)
            if xls_fileName == 'PathError':         
                pass
            else:
                if not os.path.isfile(xls_fileName):
                    # sg.popup("File not exist!") # build-in pipup window
                    MessageBox(None, xls_fileName + ' ' + lang.msg_box_file_not_exist, lang.msg_box_file_op_title, 0)
                else:
                    error_qty = excelProcessor(xls_fileName, values['exp error log'])                
                    if error_qty != 0:
                        MessageBox(None, str(error_qty) + lang.log_msg_found_error, lang.msg_box_file_op_title, 0)
                    window.Element('Result').Update(lang.gui_success)  #showing process result
        elif event == 'Last xls chkbox':
            #// TODO:待實作記錄excel檔案位置功能
            st.gen_record_last_transaction_file = values['Last xls chkbox']
            st.gen_last_transaction_file_path = values['Browse']
            # print(st.gen_record_last_transaction_file, st.gen_last_transaction_file_path)
            # print(type(st.gen_record_last_transaction_file))
            saveSetting(st)
        elif event is None or event == 'Exit':
            window.Close()
            return False

if __name__ == '__main__':
    continue_program = True
    sg.change_look_and_feel('Dark Blue 3')  # windows colorful
    MessageBox = ctypes.windll.user32.MessageBoxW
    while continue_program:
        st = loadSetting(setting_file_name='settings.json')
        lang = loadLang(st.gen_set_lang)
        window = setWindow(lang, st)
        continue_program = main(window, lang)
