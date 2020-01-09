import os
import time
import datetime
import shutil
import PySimpleGUI as sg
from openpyxl import load_workbook, utils
from openpyxl.styles import colors, PatternFill, Alignment, Font
# import openpyxl.worksheet
import ctypes
from datetime import datetime, date, time
import re
import json
import settings
import language

# openpyxl.utils.cell.coordinate_from_string('B3')  // ('B', 3)
# openpyxl.utils.cell.column_index_from_string(a[0])  // 2
# openpyxl.utils.cell.coordinate_to_tuple('B3')  // (3, 2)
# openpyxl.utils.cell.get_column_letter(3) // C

program_ver = 'Beta 1.2'

# load program setting from settings.json
def loadSetting(setting_file_name='settings.json'):
    """[load program setting from settings.json]
    
    Keyword Arguments:
        setting_file_name {str} -- [setting file you want to load] (default: {'settings.json'})
    
    Returns:
        [setting instance] {obj} -- [description]
    """
    try:
        # open json file and read
        with open(setting_file_name, 'r', encoding="utf-8") as setting:
            my_setting = json.loads(setting.read())
    except FileNotFoundError:
        MessageBox(None, setting_file_name + ' ' + lang.msg_box_file_not_exist, lang.msg_box_file_op_title, 0)
    else:
        return settings.Thfset(my_setting)
    finally:
        setting.close()

# edit program setting in settings.json
def editSetting(st, lang):
    odd_color_status = False
    even_color_status = False
    # GUI layout
    layout = [
        [sg.Text(lang.st_localization), sg.InputCombo(st.gen_ava_lang_for_GUI, size=(
            20, 1), default_value=st.gen_set_lang, key='set_lang', readonly=True)],
        [sg.Text('_' * 100, size=(55, 1))],
        [sg.Text(lang.st_xls_fmt_setting)],
        [sg.Text(lang.st_odd_col_color, size=(18, 1)), sg.InputText(st.xls_fmt_color_for_odd_column, key='odd_col_color')],
        [sg.Text(lang.st_even_col_color, size=(18, 1)), sg.InputText(st.xls_fmt_color_for_even_column, key='even_col_color')],
        [sg.Text(lang.st_disp_date_fmt, size=(18, 1)), sg.InputText(st.xls_fmt_display_date_format, key='date_fmt')],        
        [sg.Button(lang.st_ok, key='OK'), sg.Cancel(lang.st_cancel, key='Cancel'), sg.Checkbox(lang.st_backup_settings, default=st.gen_backup_setting, enable_events=True, key='backup_settings')]
    ]
    
    window = sg.Window(lang.st_setting_window_title, auto_size_text=True,
                   default_element_size=(40, 10)).Layout(layout)

    # process GUI event
    while True:
        event, values = window.Read()
        print('event: ', event, '\nvalues:', values)  # debug message
        if event == 'OK':
            # check whether any setting change or not
            if values['set_lang'] != st.gen_set_lang or values['odd_col_color'] != st.xls_fmt_color_for_odd_column or values['even_col_color'] != st.xls_fmt_color_for_even_column or values['date_fmt'] != st.xls_fmt_display_date_format:
                st.gen_set_lang = values['set_lang']
                # https://stackoverflow.com/questions/30241375/python-how-to-check-if-string-is-a-hex-color-code
                # check whether the input values satisfy the format
                if re.search(r'^(?:[0-9a-fA-F]{3}){1,2}$', values['odd_col_color']):
                    st.xls_fmt_color_for_odd_column = values['odd_col_color']
                    odd_color_status = True
                else:
                    window.Element('odd_col_color').Update('Wrong')
                    MessageBox(None, lang.msg_box_msg_odd_col_color_fmt,
                            lang.msg_box_color_fmt_wrong_title, 0)
                if re.search(r'^(?:[0-9a-fA-F]{3}){1,2}$', values['even_col_color']):
                    st.xls_fmt_color_for_even_column = values['even_col_color']
                    even_color_status = True
                else:
                    window.Element('even_col_color').Update('Wrong')
                    MessageBox(None, lang.msg_box_msg_even_col_color_fmt,
                            lang.msg_box_color_fmt_wrong_title, 0)
                # if color format ok, save settings
                if odd_color_status and even_color_status:
                    st.gen_backup_setting = values['backup_settings']
                    saveSetting(st.updateSettings(), values['backup_settings'])
                    window.close()
                    return True
            else:
                MessageBox(None, lang.msg_box_settings_file_not_change,
                           lang.msg_box_file_op_title, 0)
                window.close()
                return False
        elif event is None or event == 'Cancel':
            window.close()
            return False

# save setting to settings.json
def saveSetting(settings_obj, backup_settings=False, settings_file_name='settings.json'):
    if backup_settings:
        basename, extension = os.path.splitext(settings_file_name)
        # update setting file version info
        st.ver_info_ver += 1
        st.ver_info_date = datetime.strftime(datetime.now(), '%Y/%m/%d')
        shutil.copy(settings_file_name, basename + '_' +
                    datetime.strftime(datetime.now(), '%Y%m%d') + 'v' + str(st.ver_info_ver) + extension)
    # save settings
    with open(settings_file_name, 'w', encoding="utf-8") as settings_to_be_save:
        json.dump(settings_obj, settings_to_be_save,
                  ensure_ascii=False, indent=4)

# load langage from lang.json
def loadLang(lang_code='English (enUS)'):
    lang_file_name = 'lang_' + st.gen_set_lang[-5:-1] + '.json'
    try:
        # open json file and read
        with open(lang_file_name, 'r', encoding="utf-8") as lang:
            lang_setting = json.loads(lang.read())
    except FileNotFoundError:
        MessageBox(None, lang_file_name + ' ' + lang.msg_box_file_not_exist, lang.msg_box_file_op_title, 0)
    else:
        return language.Lang(lang_setting)
    finally:
        lang.close()

# setup window layout
def setWindow(lang, st):
    # setup window layout
    layout = [[sg.Text(lang.gui_program_setting + ':')],
              [sg.Text(lang.gui_file + ':', justification='right'),
               sg.Text('', size=(65, 1), key='settings status')],
              [sg.FileBrowse(lang.gui_load_setting_file, target='settings status', file_types=((lang.gui_settings_file, "*settings*.json"),), key='Load Setting File'),
               sg.Button(lang.gui_apply_settings, key='Apply Settings'), sg.Button(lang.gui_open_setting_editor, key='Open Setting Editor')],
              [sg.Text('_' * 100, size=(70, 1))],
              [sg.Text(lang.gui_load_trade_history_file + ':')],
              [sg.Text(lang.gui_file + ':', justification='right'),
               sg.Text('', size=(65, 1), key='it_filePath')],
              [sg.FileBrowse(lang.gui_load_trade_history_file, target='it_filePath', file_types=((lang.gui_spreadsheet_files, "*.xls"),
                                                                                                 (lang.gui_spreadsheet_files, "*.xlsx"),), key='Browse'),
               sg.Button(lang.gui_process_history, key='Process History'),
               sg.Checkbox(lang.gui_exp_error_log, default=st.gen_exp_error_log, enable_events=True, key='exp error log')],
              [sg.ProgressBar(100, orientation='h', size=(50,15), auto_size_text=True, key='process status')],
              [sg.Text(lang.gui_result + ':'),
               sg.Text('', size=(20, 1), key='Result')],
              [sg.Text('_' * 100, size=(70, 1))],
              [sg.Text(lang.gui_ver + ': ' + program_ver, size=(80, 1), font='Arial 8'),
               sg.Button(lang.gui_exit, size=(5, 1), key='Exit')]]

    # rendering window
    window = sg.Window(lang.gui_title, auto_size_text=True, default_element_size=(
        40, 10), resizable=False).Layout(layout)
    return window

# get excel file name to be processed
def getXlsFileName(filePath, lang):
    if filePath == None or filePath == '':        
        MessageBox(None, lang.msg_box_select_file_first, lang.msg_box_file_op_title, 0)
        return 'PathError'                
    else:
        xls_fileName = os.path.basename(filePath)        
        os.chdir(os.path.dirname(filePath))    
        return xls_fileName

# excel processor
def excelProcessor(xls_fileName, exp_error_log, st, lang, window, symbol_list = []):
    # sheet_list = ['Sorted trade history','ORDINARY DIVIDEND','W-8 WITHHOLDING','WIRING INFO','Ver','log']
    sheet_list = lang.xls_sheet_names
    error_log_qty = 0
    if symbol_list == []:
        gather_symbol = True
    else:
        gather_symbol = False
    # loading workbook
    wb = load_workbook(xls_fileName)
    # create sheets and setting status
    for i in range(len(sheet_list)):
        if not sheet_list[i] in wb.sheetnames:
            wb.create_sheet(sheet_list[i])
    
    # check transactions file status
    if 'transactions(DONE)' in wb.sheetnames:
        return -2
    elif not 'transactions' in wb.sheetnames:
        return -1
    else:
        ws_tran = wb["transactions"]
    ws_STH = wb[sheet_list[0]]
    ws_OD = wb[sheet_list[1]]
    ws_WH = wb[sheet_list[2]]
    ws_WI = wb[sheet_list[3]]
    ws_ver = wb[sheet_list[4]]
    ws_log = wb[sheet_list[5]]
    ws_status = wb[sheet_list[6]]
    ws_status.protection.sheet = True
    ws_status.sheet_state = 'hidden'

    # setting excel sheets layout
    title_date = lang.xls_tt_date
    ws_STH['A2'] = title_date # Sorted trading history
    ws_OD['A1'] = title_date # DIVIDEND
    ws_WH['A1'] = title_date # WITHHOLD
    ws_WI['A1'] = title_date # WIRING INFO
    ws_WI['B1'] = lang.xls_tt_amount
    ws_ver['A1'] = title_date # Ver
    ws_ver['B1'] = lang.xls_tt_ver
    ws_log['A1'] = lang.xls_tt_event
    ws_log['B1'] = lang.xls_tt_msg
    ws_status['A1'] = lang.xls_stat_sym_list
    ws_status['A4'] = lang.xls_stat_sym_qty
    ws_status['A7'] = 'INTERNAL TRANSFER BETWEEN ACCOUNTS OR ACCOUNT TYPES'
    ws_status['A10'] = 'QUALIFIED DIVIDEND'
    ws_status['A13'] = 'CASH IN LIEU'
    ws_status['A16'] = 'DIVIDEND SHORT SALE'
    ws_status['A19'] = 'NON-TAXABLE DIVIDENDS'
    ws_status['A22'] = 'MANDATORY - EXCHANGE'

    # start sheets process
    iter_date_STH = ''
    iter_date_OD = ''
    iter_date_WH = ''
    # iter_date_WI = ''
    # check file status before process
    if ws_status.cell(1, 8).value == None or ws_status.cell(1, 8).value == 'False':
        ws_STH_have_inter_trans = False
    elif ws_status.cell(1, 8).value == 'True':
        ws_STH_have_inter_trans = True
    if ws_status['A23'].value == None or ws_status['A23'].value == 'False':
        ws_STH_have_mandi_exchange = False
    elif ws_status['A23'].value == 'True':
        ws_STH_have_mandi_exchange = True
    if ws_status.cell(1, 11).value == None or ws_status.cell(1, 11).value == 'False':
        ws_OD_have_quali_div = False
    elif ws_status.cell(1, 11).value == 'True':
        ws_OD_have_quali_div = True
    if ws_status.cell(1, 14).value == None or ws_status.cell(1, 14).value == 'False':
        ws_OD_have_cashInLieu = False
    elif ws_status.cell(1, 14).value == 'True':
        ws_OD_have_cashInLieu = True
    if ws_status['A17'].value == None or ws_status['A17'].value == 'False':
        ws_OD_have_div_short = False
    elif ws_status['A17'].value == 'True':
        ws_OD_have_div_short = True
    if ws_status['A20'].value == None or ws_status['A20'].value == 'False':
        ws_OD_have_nontax_div = False
    elif ws_status['A20'].value == 'True':
        ws_OD_have_nontax_div = True
    
    # retrieve symbol list from file
    if ws_status['A2'].value != None:
        existing_symbol_list = [sym.strip() for sym in ws_status['A2'].value.split(',')]
        print(existing_symbol_list)
        if gather_symbol:
            symbol_list = existing_symbol_list
    
    # for future feature.  If user know all stock symbol for this batch trading history, they could input it on GUI.
    if not gather_symbol:
        symbol_be_handled = []
        for symbol in symbol_list:
            if not symbol in existing_symbol_list:
                existing_symbol_list.append(symbol)
                symbol_be_handled.append(symbol)
        symbol_list = existing_symbol_list
        for tr_symbol in symbol_be_handled:
            symbol_index = symbol_list.index(tr_symbol)  # stock symbol
            ws_STH.cell(1, symbol_index*4+2).value = tr_symbol
            ws_STH.merge_cells(start_row=1, start_column=symbol_index *
                            4+2, end_row=1, end_column=symbol_index*4+5)  # merge cell
            ws_STH.cell(1, symbol_index*4+2).alignment = Alignment(
                horizontal="center", vertical="center")  # centering text
            ws_STH.cell(2, symbol_index*4+2).value = lang.xls_tt_quantity
            ws_STH.cell(2, symbol_index*4+3).value = lang.xls_tt_price
            ws_STH.cell(2, symbol_index*4+4).value = lang.xls_tt_commission
            ws_STH.cell(2, symbol_index*4+5).value = lang.xls_tt_amount
            ws_OD.cell(1, symbol_index+2).value = tr_symbol
            ws_OD.cell(1, symbol_index+2).alignment = Alignment(
                horizontal="center", vertical="center")  # centering text
            ws_WH.cell(1, symbol_index+2).value = tr_symbol
            ws_WH.cell(1, symbol_index+2).alignment = Alignment(
                horizontal="center", vertical="center")  # centering text

    for i in range(2, ws_tran.max_row):
        tr_date = ws_tran.cell(i, 1)
        # date process
        temp_date_for_sheet = datetime.strptime(tr_date.value, "%m/%d/%Y")
        date_for_sheet = temp_date_for_sheet.strftime(st.xls_fmt_display_date_format)
        tr_description = ws_tran.cell(i, 3)
        tr_qty = ws_tran.cell(i, 4)
        tr_symbol = ws_tran.cell(i, 5)
        tr_price = ws_tran.cell(i, 6)
        tr_fee = ws_tran.cell(i, 7)
        tr_amount = ws_tran.cell(i, 8)
        # processing sheets format by stock symbols and gather all stock symbol from 'E5'
        if gather_symbol:
            if not tr_symbol.value in symbol_list and tr_symbol.value != None:
                symbol_list.append(tr_symbol.value)  # gather stock symbol
                symbol_index = symbol_list.index(tr_symbol.value)  # stock symbol
                ws_STH.cell(1, symbol_index*4+2).value = tr_symbol.value
                ws_STH.merge_cells(start_row=1, start_column=symbol_index *
                                4+2, end_row=1, end_column=symbol_index*4+5)  # merge cell
                ws_STH.cell(1, symbol_index*4+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                ws_STH.cell(2, symbol_index*4+2).value = lang.xls_tt_quantity
                ws_STH.cell(2, symbol_index*4+3).value = lang.xls_tt_price
                ws_STH.cell(2, symbol_index*4+4).value = lang.xls_tt_commission
                ws_STH.cell(2, symbol_index*4+5).value = lang.xls_tt_amount
                ws_OD.cell(1, symbol_index+2).value = tr_symbol.value
                ws_OD.cell(1, symbol_index+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text
                ws_WH.cell(1, symbol_index+2).value = tr_symbol.value
                ws_WH.cell(1, symbol_index+2).alignment = Alignment(
                    horizontal="center", vertical="center")  # centering text

        # sorting trading history by the description of transactions
        # cash flows
        if 'WIRE INCOMING' in tr_description.value:
            ws_WI.insert_rows(2)  # add new row
            ws_WI.cell(2, 1).value = date_for_sheet  # date
            ws_WI.cell(2, 2).value = tr_amount.value  # amount
        elif 'CLIENT REQUESTED ELECTRONIC FUNDING DISBURSEMENT' in tr_description.value:
            ws_WI.insert_rows(2)  # add new row
            ws_WI.cell(2, 1).value = date_for_sheet  # date
            ws_WI.cell(2, 2).value = tr_amount.value  # amount
            if ws_WI.cell(1, 3).value == None:
                ws_WI.cell(1, 3).value = lang.xls_tt_remark
            ws_WI.cell(2, 3).value = lang.xls_msg_client_req_e_funding_dist
        elif 'CLIENT REQUESTED ELECTRONIC FUNDING RECEIPT' in tr_description.value:
            ws_WI.insert_rows(2)  # add new row
            ws_WI.cell(2, 1).value = date_for_sheet  # date
            ws_WI.cell(2, 2).value = tr_amount.value  # amount
            if ws_WI.cell(1, 3).value == None:
                ws_WI.cell(1, 3).value = lang.xls_tt_remark
            ws_WI.cell(2, 3).value = lang.xls_msg_client_req_e_funding_rec
        elif 'INTRA-ACCOUNT TRANSFER' in tr_description.value:
            ws_WI.insert_rows(2)  # add new row
            ws_WI.cell(2, 1).value = date_for_sheet  # date
            ws_WI.cell(2, 2).value = tr_amount.value  # amount
            if ws_WI.cell(1, 3).value == None:
                ws_WI.cell(1, 3).value = lang.xls_tt_remark
            ws_WI.cell(2, 3).value = lang.xls_msg_intra_account_transfer
        elif 'REBATE' in tr_description.value:
            ws_WI.insert_rows(2)  # add new row
            ws_WI.cell(2, 1).value = date_for_sheet  # date
            ws_WI.cell(2, 2).value = tr_amount.value  # amount
            if ws_WI.cell(1, 3).value == None:
                ws_WI.cell(1, 3).value = lang.xls_tt_remark
            ws_WI.cell(2, 3).value = lang.xls_msg_rebate
        # transaction
        elif 'Bought' in tr_description.value or 'Sold' in tr_description.value:
            if tr_symbol.value in symbol_list:
                symbol_index = symbol_list.index(tr_symbol.value) # get index value in list
                if tr_date.value != iter_date_STH or ws_STH.cell(3, symbol_index*4 + 2).value != None:
                    ws_STH.insert_rows(3)  # add new row
                    ws_STH.cell(3, 1).value = date_for_sheet  # date
                    iter_date_STH = tr_date.value
                ws_STH.cell(3, symbol_index*4+2).value = tr_qty.value
                ws_STH.cell(3, symbol_index*4+3).value = tr_price.value
                ws_STH.cell(3, symbol_index*4+4).value = tr_fee.value
                ws_STH.cell(3, symbol_index*4+5).value = tr_amount.value
            else: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = lang.log_evt_transaction_symbol_missing
                    temp_msg = lang.log_msg_transaction_symbol_missing
                    ws_log.cell(2, 2).value = (temp_msg.replace('-symbol-', tr_symbol.value)).replace('-xx-', str(i))
                    error_log_qty += 1
        elif 'INTERNAL TRANSFER BETWEEN ACCOUNTS OR ACCOUNT TYPES' in tr_description.value:
            if tr_symbol.value in symbol_list:
                symbol_index = symbol_list.index(tr_symbol.value) # get index value in list
                if tr_date.value != iter_date_STH or ws_STH.cell(3, symbol_index*4 + 2).value != None:
                    ws_STH.insert_rows(3)  # add new row
                    ws_STH.cell(3, 1).value = date_for_sheet  # date
                    iter_date_STH = tr_date.value
                ws_STH.cell(3, symbol_index*4+2).value = tr_qty.value
                ws_STH.cell(3, symbol_index*4+3).value = tr_price.value
                ws_STH.cell(3, symbol_index*4+4).value = tr_fee.value
                ws_STH.cell(3, symbol_index*4+5).value = tr_amount.value
                # change to bold to indicate INTERNAL TRANSFER
                ws_STH.cell(3, symbol_index*4+2).font = Font(bold=True)
                ws_STH.cell(3, symbol_index*4+3).font = Font(bold=True)
                ws_STH.cell(3, symbol_index*4+4).font = Font(bold=True)
                ws_STH.cell(3, symbol_index*4+5).font = Font(bold=True)
                ws_STH_have_inter_trans = True
            else: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = lang.log_evt_transaction_symbol_missing
                    temp_msg = lang.log_msg_transaction_symbol_missing
                    ws_log.cell(2, 2).value = (temp_msg.replace('-symbol-', tr_symbol.value)).replace('-xx-', str(i))
                    error_log_qty += 1
        elif 'MANDATORY - EXCHANGE' in tr_description.value:
            if tr_symbol.value in symbol_list:
                symbol_index = symbol_list.index(tr_symbol.value) # get index value in list
                if tr_date.value != iter_date_STH or ws_STH.cell(3, symbol_index*4 + 2).value != None:
                    ws_STH.insert_rows(3)  # add new row
                    ws_STH.cell(3, 1).value = date_for_sheet  # date
                    iter_date_STH = tr_date.value
                ws_STH.cell(3, symbol_index*4+2).value = tr_qty.value
                ws_STH.cell(3, symbol_index*4+3).value = tr_price.value
                ws_STH.cell(3, symbol_index*4+4).value = tr_fee.value
                ws_STH.cell(3, symbol_index*4+5).value = tr_amount.value
                # change to italic to indicate MANDATORY - EXCHANGE
                ws_STH.cell(3, symbol_index*4+2).font = Font(italic=True)
                ws_STH.cell(3, symbol_index*4+3).font = Font(italic=True)
                ws_STH.cell(3, symbol_index*4+4).font = Font(italic=True)
                ws_STH.cell(3, symbol_index*4+5).font = Font(italic=True)
                ws_STH_have_mandi_exchange = True
            else: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = lang.log_evt_transaction_symbol_missing
                    temp_msg = lang.log_msg_transaction_symbol_missing
                    ws_log.cell(2, 2).value = (temp_msg.replace('-symbol-', tr_symbol.value)).replace('-xx-', str(i))
                    error_log_qty += 1
        # dividend
        elif 'ORDINARY DIVIDEND' in tr_description.value:
            symbol_index = symbol_list.index(tr_symbol.value) # get index value in list
            if tr_date.value != iter_date_OD:
                ws_OD.insert_rows(2)  # add new row
                ws_OD.cell(2, 1).value = date_for_sheet  # date
                iter_date_OD = tr_date.value
            ws_OD.cell(2, symbol_index+2).value = tr_amount.value
        elif 'QUALIFIED DIVIDEND' in tr_description.value:
            ws_OD_have_quali_div = True
            if tr_date.value != iter_date_OD:
                ws_OD.insert_rows(2)  # add new row
                ws_OD.cell(2, 1).value = date_for_sheet  # date
                iter_date_OD = tr_date.value
            ws_OD.cell(2, symbol_index+2).value = tr_amount.value
            ws_OD.cell(2, symbol_index+2).font = Font(italic=True)
        elif 'CASH IN LIEU' in tr_description.value:
            ws_OD_have_cashInLieu = True
            if tr_date.value != iter_date_OD:
                ws_OD.insert_rows(2)  # add new row
                ws_OD.cell(2, 1).value = date_for_sheet  # date
                iter_date_OD = tr_date.value
            ws_OD.cell(2, symbol_index+2).value = tr_amount.value
            ws_OD.cell(2, symbol_index+2).font = Font(bold=True)
        elif 'DIVIDEND SHORT SALE' in tr_description.value:
            ws_OD_have_div_short = True
            if tr_date.value != iter_date_OD:
                ws_OD.insert_rows(2)  # add new row
                ws_OD.cell(2, 1).value = date_for_sheet  # date
                iter_date_OD = tr_date.value
            ws_OD.cell(2, symbol_index+2).value = tr_amount.value
            ws_OD.cell(2, symbol_index+2).font = Font(color=colors.RED)
        elif 'NON-TAXABLE DIVIDENDS' in tr_description.value:
            ws_OD_have_nontax_div = True
            if tr_date.value != iter_date_OD:
                ws_OD.insert_rows(2)  # add new row
                ws_OD.cell(2, 1).value = date_for_sheet  # date
                iter_date_OD = tr_date.value
            ws_OD.cell(2, symbol_index+2).value = tr_amount.value
            ws_OD.cell(2, symbol_index+2).font = Font(color=colors.BLUE)
        # withholding
        elif 'WITHHOLDING' in tr_description.value:            
            if tr_symbol.value == None: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = 'WITHHOLDING'
                    ws_log.cell(2, 2).value = lang.log_msg_withholding_symbol_missing.replace('-xx-', str(i))
                    error_log_qty += 1
            else:
                if tr_date.value != iter_date_WH:
                    ws_WH.insert_rows(2)  # add new row
                    ws_WH.cell(2, 1).value = date_for_sheet  # date
                    iter_date_WH = tr_date.value
                ws_WH.cell(2, symbol_index+2).value = tr_amount.value
        elif 'FOREIGN TAX WITHHELD' in tr_description.value:            
            if tr_symbol.value == None: # export error message
                if exp_error_log:
                    ws_log.insert_rows(2)
                    ws_log.cell(2, 1).value = 'FOREIGN TAX WITHHELD'
                    ws_log.cell(2, 2).value = lang.log_msg_withholding_symbol_missing.replace('-xx-', str(i))
                    error_log_qty += 1
            else:
                if tr_date.value != iter_date_WH:
                    ws_WH.insert_rows(2)  # add new row
                    ws_WH.cell(2, 1).value = date_for_sheet  # date
                    iter_date_WH = tr_date.value
                ws_WH.cell(2, symbol_index+2).value = tr_amount.value
        # skip events
        elif any([keyword in tr_description.value for keyword in st.xls_skip_event]):
            if exp_error_log:
                ws_log.insert_rows(2)
                ws_log.cell(2, 1).value = lang.log_evt_event_skip
                ws_log.cell(2, 2).value = (lang.log_msg_event_skip.replace('-description-', tr_description.value)).replace('-xx-', str(i))
        # export error message
        else:
            if exp_error_log:
                ws_log.insert_rows(2)
                ws_log.cell(2, 1).value = lang.log_evt_description_keyword_missing
                ws_log.cell(2, 2).value = (lang.log_msg_description_keyword_missing.replace('-description-', tr_description.value)).replace('-xx-', str(i))
                error_log_qty += 1
        
        # update progress bar status
        window.Element('process status').UpdateBar((i+1)/ws_tran.max_row*80)

    # record symbol list and quantity for this batch of trading history
    ws_status['A2'] = ','.join(symbol_list)
    ws_status['A5'] = len(symbol_list)
    
    # create sheet for excel remark
    opt_sheet_list = lang.xls_opt_sheet_names
    if any([ws_STH_have_inter_trans, ws_STH_have_mandi_exchange, ws_OD_have_quali_div, ws_OD_have_cashInLieu, ws_OD_have_div_short, ws_OD_have_nontax_div]) and not opt_sheet_list[0] in wb.sheetnames:
        wb.create_sheet(opt_sheet_list[0], index=5)
        ws_remark = wb[opt_sheet_list[0]]
        ws_remark['A1'] = sheet_list[0]
        for i in range(1,5):
            ws_remark.cell(1, i).fill= PatternFill(fgColor=st.xls_fmt_color_for_even_column, fill_type="solid")
        ws_remark['A5'] = sheet_list[1]
        for i in range(1,5):
            ws_remark.cell(5, i).fill= PatternFill(fgColor=st.xls_fmt_color_for_odd_column, fill_type="solid")
        ws_remark.protection.sheet = True
    if ws_STH_have_inter_trans:
        ws_remark['A2'] = lang.xls_msg_bold_indication_ITA
        ws_remark['A2'].font = Font(bold=True)
        ws_status['A8'] = 'True'
    if ws_STH_have_mandi_exchange:
        ws_remark['A3'] = lang.xls_msg_italic_mandi_exchange
        ws_remark['A3'].font = Font(italic=True)
        ws_status['A23'] = 'True'
    if ws_OD_have_quali_div:
        ws_remark['A6'] = lang.xls_msg_italic_qual_div
        ws_remark['A6'].font = Font(italic=True)
        ws_status['A11'] = 'True'
    if ws_OD_have_cashInLieu:
        ws_remark['A7'] = lang.xls_msg_bold_cashInLieu
        ws_remark['A7'].font = Font(bold=True)
        ws_status['A14'] = 'True'
    if ws_OD_have_div_short:
        ws_remark['A8'] = lang.xls_msg_red_div_short
        ws_remark['A8'].font = Font(color=colors.RED)
        ws_status['A17'] = 'True'
    if ws_OD_have_nontax_div:
        ws_remark['A9'] = lang.xls_msg_blue_nontax_div
        ws_remark['A9'].font = Font(color=colors.BLUE)
        ws_status['A20'] = 'True'
    
    ws_tran.title = 'transactions(DONE)'

    #// TODO: performance optimization required: need a new approach to set the cell color
    # setting cell color
    for k in range(len(symbol_list)):
        if k % 2 == 0:
            color_fill = st.xls_fmt_color_for_even_column
        else:
            color_fill = st.xls_fmt_color_for_odd_column
        for row in ws_STH.iter_rows(min_col=k*4+2, min_row=1, max_col=k*4+5, max_row=ws_STH.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor=color_fill, fill_type="solid")
        for row in ws_OD.iter_rows(min_col=k+2, min_row=1, max_col=k+2, max_row=ws_OD.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor=color_fill, fill_type="solid")
        for row in ws_WH.iter_rows(min_col=k+2, min_row=1, max_col=k+2, max_row=ws_WH.max_row):
            for cell in row:
                cell.fill = PatternFill(fgColor=color_fill, fill_type="solid")
        # update progress bar status
        window.Element('process status').UpdateBar(80+(k+1)/len(symbol_list)*19)
    
    # version control
    file_version = ws_ver['B2'].value  # get current file version
    [file, ext] = os.path.splitext(xls_fileName)
    if file_version == None:
        ws_ver['A2'] = date.today().strftime("%Y/%m/%d")  # date
        ws_ver['B2'] = 0
        file_version = 0
        fileNameRev = file + '_r' + str(file_version) + ext
    else:
        ws_ver.insert_rows(2)  # add new row
        ws_ver['A2'] = date.today().strftime("%Y/%m/%d")  # date
        file_version += 1  # update version number
        ws_ver['B2'] = file_version
        if '_r' in file:
            fileNameRev = file[:file.find('_r')+2] + str(file_version) + ext
        else:
            fileNameRev = file + '_r' + str(file_version) + ext

    wb.save(fileNameRev)  # save processed file
    wb.close()
    return error_log_qty
    
# Main Program
def main(window, st, lang):
    xls_fileName = None
    while True:
        event, values = window.Read()
        print('event: ', event, '\nvalues:', values)  # debug message
        if event == 'Apply Settings':            
            if values['Load Setting File'] == None or values['Load Setting File'] == '':
                MessageBox(None, lang.msg_box_select_file_first, lang.msg_box_file_op_title, 0)
            else:
                window.Close()
                return True, os.path.basename(values['Load Setting File'])
        elif event == 'Open Setting Editor':
            if editSetting(st, lang):
                window.close()
                return True, 'settings.json'
        elif event == 'Process History':
            xls_fileName = getXlsFileName(values['Browse'], lang)
            if xls_fileName == 'PathError':
                pass
            else:
                if not os.path.isfile(xls_fileName):
                    # sg.popup("File not exist!") # build-in pipup window
                    MessageBox(None, xls_fileName + ' ' + lang.msg_box_file_not_exist, lang.msg_box_file_op_title, 0)
                else:
                    error_qty = excelProcessor(xls_fileName, values['exp error log'], st, lang, window, [])
                    if error_qty > 0:
                        MessageBox(None, lang.log_msg_found_error.replace('-xx-', str(error_qty)), lang.msg_box_file_op_title, 0)
                        window.Element('it_filePath').Update('')
                        window.Element('Result').Update(lang.gui_success)  #showing process result
                        # update progress bar status
                        window.Element('process status').UpdateBar(100)
                    elif error_qty == 0:
                        window.Element('it_filePath').Update('')
                        window.Element('Result').Update(lang.gui_success)  #showing process result
                        # update progress bar status
                        window.Element('process status').UpdateBar(100)
                    elif error_qty == -1:
                        MessageBox(None, lang.msg_box_trading_sht_not_exist, lang.msg_box_file_op_title, 0)
                        window.Element('Result').Update(lang.msg_box_trading_sht_not_exist)  #showing error message
                    elif error_qty == -2:
                        MessageBox(None, lang.msg_box_file_processed, lang.msg_box_file_op_title, 0)
                        window.Element('Result').Update(lang.msg_box_file_processed)  #showing error message
        elif event is None or event == 'Exit':
            window.Close()
            return False, 'settings.json'

if __name__ == '__main__':
    continue_program = True
    sg.change_look_and_feel('Dark Blue 3')  # windows colorful
    MessageBox = ctypes.windll.user32.MessageBoxW
    setting_file_name = 'settings.json'
    while continue_program:
        st = loadSetting(setting_file_name)
        lang = loadLang(st.gen_set_lang)
        window = setWindow(lang, st)
        # if setting have any changes, program will restart automatically
        continue_program, setting_file_name = main(window, st, lang)
